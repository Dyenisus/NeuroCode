"""
EEG Artifact Submission Grader

Scores one team CSV against the hardcoded answer key.
Submission rows are matched to answer-key rows by best IoU — order does not matter.

Usage:
    python submission_check.py path/to/group-id_group-name.csv
    python submission_check.py path/to/group-id_group-name.csv --rows rows.csv --summary summary.csv
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from statistics import mean
from openpyxl import load_workbook

# ============================================================
# ANSWER KEY
# Edit these rows to match the official workshop annotations.
# ============================================================

ANSWER_KEY = [
    {
        "artifact_category": "medium_task_block",
        "artifact_class":    "swallow",
        "start_time":        606.279,
        "end_time":          608.278,
    },
    {
        "artifact_category": "medium_task_block",
        "artifact_class":    "tongue",
        "start_time":        808.574,
        "end_time":          813.574,
    },
    {
        "artifact_category": "long_task_block",
        "artifact_class":    "chew",
        "start_time":        737.544,
        "end_time":          749.562,
    },
    {
        "artifact_category": "long_task_block",
        "artifact_class":    "eyebrow",
        "start_time":        869.285,
        "end_time":          870.270,
    },
    {
        "artifact_category": "long_task_block",
        "artifact_class":    "hor_headm",
        "start_time":        468.858,
        "end_time":          474.858,
    },
    {
        "artifact_category": "long_task_block",
        "artifact_class":    "ver_headm",
        "start_time":        528.876,
        "end_time":          534.874,
    },
    {
        "artifact_category": "macroscopic_state",
        "artifact_class":    "blink",
        "start_time":        379.701,
        "end_time":          440.707,
    },
    {
        "artifact_category": "long_task_block",
        "artifact_class":    "ver_headm",
        "start_time":        1152.872,
        "end_time":          1158.856,
    },
    {
        "artifact_category": "long_task_block",
        "artifact_class":    "hor_headm",
        "start_time":        1146.856,
        "end_time":          1152.857,
    }
]

# ============================================================
# CATEGORY RULES
# onset_tolerance  — max allowed |pred_start - true_start| in seconds
# offset_tolerance — max allowed |pred_end   - true_end|   in seconds
# ============================================================

CATEGORY_RULES = {
    "micro_event": {
        "classes":           ["ver_eyem"],
        "onset_tolerance":   0.20,   # ± 200 ms
        "offset_tolerance":  0.30,   # ± 300 ms
    },
    "medium_task_block": {
        "classes":           ["tongue", "swallow"],
        "onset_tolerance":   0.50,   # ± 500 ms
        "offset_tolerance":  1.00,   # ± 1000 ms
    },
    "long_task_block": {
        "classes":           ["hor_headm", "ver_headm", "eyebrow", "chew"],
        "onset_tolerance":   1.50,   # ± 1.5 s
        "offset_tolerance":  2.00,   # ± 2.0 s
    },
    "macroscopic_state": {
        "classes":           ["blink", "close_base", "open_base"],
        "onset_tolerance":   3.00,   # ± 3.0 s
        "offset_tolerance":  3.00,   # ± 3.0 s
    },
}

# ============================================================
# POINT WEIGHTS
# Change these values to adjust competition scoring.
# ============================================================

MAX_INTERVAL_POINTS = 10.0   # Scales with IoU once timing passes

CATEGORY_POINTS = {
    "micro_event":       20.0,
    "medium_task_block": 12.5,
    "long_task_block":  10.0,
    "macroscopic_state": 5.0,
}

# ============================================================
# TOLERANCE PAD
# Multiplicative padding added on top of category onset/offset tolerances.
# e.g. 0.10 means each tolerance is widened by 10% (tol * 1.10).
# ============================================================

TOLERANCE_PAD = 0.10   # fraction (0.10 = +10%); must be >= 0

if TOLERANCE_PAD < 0:
    raise ValueError(f"TOLERANCE_PAD must be >= 0, got {TOLERANCE_PAD}")

# ============================================================
# REQUIRED COLUMNS (do not change unless the spec changes)
# ============================================================

REQUIRED_COLUMNS = ["artifact_category", "artifact_class", "start_time", "end_time"]


# ────────────────────────────────────────────────────────────
# I/O helpers
# ────────────────────────────────────────────────────────────

def _norm(name: str) -> str:
    """Normalize a column header to lowercase_with_underscores."""
    return name.strip().lower().replace(" ", "_")


def _validate_columns(headers: list[str]) -> dict[str, str]:
    col_map = {_norm(h): h for h in headers}
    missing = [c for c in REQUIRED_COLUMNS if c not in col_map]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")
    return col_map


def read_submission(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames is None:
                return []
            col_map = _validate_columns(list(reader.fieldnames))
            return [
                {c: raw[col_map[c]].strip() for c in REQUIRED_COLUMNS}
                for raw in reader
            ]

    if suffix == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            return []
        headers = [("" if v is None else str(v)) for v in header_row]
        col_map = _validate_columns(headers)
        header_index = {h: i for i, h in enumerate(headers)}
        rows = []
        for raw in row_iter:
            rows.append({
                c: ("" if (v := raw[header_index[col_map[c]]]) is None else str(v).strip())
                for c in REQUIRED_COLUMNS
            })
        return rows

    raise ValueError(f"Unsupported file type '{path.suffix}'. Use .csv or .xlsx.")


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


# ────────────────────────────────────────────────────────────
# Scoring helpers
# ────────────────────────────────────────────────────────────

def iou(p_start: float, p_end: float, t_start: float, t_end: float) -> float:
    intersection = max(0.0, min(p_end, t_end) - max(p_start, t_start))
    union = max(p_end, t_end) - min(p_start, t_start)
    return intersection / union if union > 0 else 0.0


def is_blank(row: dict[str, str] | None) -> bool:
    return row is None or all(row.get(c, "").strip() == "" for c in REQUIRED_COLUMNS)


def parse_times(row: dict[str, str]) -> tuple[bool, str, float, float]:
    """
    Parse and validate start/end times from a submission row.
    Returns (is_valid, status_tag, pred_start, pred_end).
    """
    try:
        p_start = float(row["start_time"])
        p_end   = float(row["end_time"])
    except ValueError:
        return False, "non_numeric_time", 0.0, 0.0

    if p_end <= p_start:
        return False, "end_not_after_start", p_start, p_end

    return True, "ok", p_start, p_end


# ────────────────────────────────────────────────────────────
# IoU-based matching
# ────────────────────────────────────────────────────────────

def match_predictions(
    submission: list[dict[str, str]],
) -> tuple[dict[int, tuple[dict[str, str], float]], list[dict[str, str]]]:
    """
    Match each valid submission row to its best-IoU answer-key row.
    Each submission row and each answer-key row is used at most once.
    Higher-IoU pairs are assigned before lower-IoU pairs.

    Returns a tuple of:
      - assignments: dict mapping answer_key_index -> (matched submission row, iou_score).
      - invalid_preds: list of non-blank rows whose times could not be parsed,
        in submission order. These are surfaced so the caller can stamp the
        real parse-time status instead of "unmatched".
    """
    # Collect all valid submission rows with their parsed times.
    # Also collect non-blank rows that failed parse_times so the caller can
    # report the real status (e.g. non_numeric_time, end_not_after_start)
    # instead of silently turning them into "unmatched" truth rows.
    valid_preds: list[tuple[int, dict[str, str], float, float]] = []
    invalid_preds: list[dict[str, str]] = []
    for sub_idx, row in enumerate(submission):
        if is_blank(row):
            continue
        is_valid, _, p_start, p_end = parse_times(row)
        if is_valid:
            valid_preds.append((sub_idx, row, p_start, p_end))
        else:
            invalid_preds.append(row)

    # Build all (iou_score, answer_key_index, pred_entry) candidates
    candidates: list[tuple[float, int, tuple]] = []
    for key_idx, truth in enumerate(ANSWER_KEY):
        t_start = float(truth["start_time"])
        t_end   = float(truth["end_time"])
        for sub_idx, row, p_start, p_end in valid_preds:
            score = iou(p_start, p_end, t_start, t_end)
            if score >= 0:
                candidates.append((score, key_idx, (sub_idx, row)))

    # Greedy assignment: best IoU first, each side used at most once
    candidates.sort(key=lambda x: x[0], reverse=True)
    matched_keys: set[int] = set()
    matched_subs: set[int] = set()
    assignments: dict[int, tuple[dict[str, str], float]] = {}

    for score, key_idx, (sub_idx, row) in candidates:
        if key_idx in matched_keys or sub_idx in matched_subs:
            continue
        assignments[key_idx] = (row, score)
        matched_keys.add(key_idx)
        matched_subs.add(sub_idx)

    return assignments, invalid_preds


# ────────────────────────────────────────────────────────────
# Row scoring
# ────────────────────────────────────────────────────────────

def score_row(row_number: int, truth: dict, pred: dict[str, str] | None, overlap: float) -> dict:
    result = {
        "row_id":            row_number,
        "points_earned":     0.0,
        "interval_accuracy": 0.0,
        "category_accuracy": 0,
        "artifact_accuracy": 0,
        "status":            "unmatched",
    }

    if pred is None:
        return result

    t_cat   = truth["artifact_category"]
    t_class = truth["artifact_class"]
    rules   = CATEGORY_RULES[t_cat]

    is_valid, status, p_start, p_end = parse_times(pred)
    result["status"] = status
    if not is_valid:
        return result

    cat_correct   = pred["artifact_category"] == t_cat
    class_correct = pred["artifact_class"]    == t_class

    # Accuracy is recorded regardless of timing tolerance
    result["category_accuracy"] = int(cat_correct)
    result["artifact_accuracy"] = int(class_correct)

    # Timing tolerance check using truth row's tolerances
    t_start   = float(truth["start_time"])
    t_end     = float(truth["end_time"])
    onset_tol  = rules["onset_tolerance"]  * (1 + TOLERANCE_PAD)
    offset_tol = rules["offset_tolerance"] * (1 + TOLERANCE_PAD)
    onset_ok   = abs(p_start - t_start) <= onset_tol
    offset_ok  = abs(p_end   - t_end)   <= offset_tol

    if not (onset_ok and offset_ok):
        result["status"] = "interval_outside_tolerance"
        return result

    # Timing passed — compute points using the already-computed overlap
    interval_pts = MAX_INTERVAL_POINTS * overlap
    category_pts = CATEGORY_POINTS.get(t_cat, 0.0)   if cat_correct   else 0.0

    result["points_earned"]     = round(interval_pts + category_pts, 2)
    result["interval_accuracy"] = round(overlap, 4)
    result["status"]            = status
    return result


# ────────────────────────────────────────────────────────────
# Submission scoring
# ────────────────────────────────────────────────────────────

def score_submission(path: Path) -> tuple[list[dict], dict]:
    submission  = read_submission(path)
    assignments, invalid_preds = match_predictions(submission)

    # Consume invalid predictions from the front so that unmatched truth rows
    # surface the real parse_times status (e.g. non_numeric_time) instead of
    # silently being stamped "unmatched".
    remaining_invalid = list(invalid_preds)
    row_results = []
    for key_idx, truth in enumerate(ANSWER_KEY):
        matched = assignments.get(key_idx)
        if matched:
            pred, overlap = matched
        elif remaining_invalid:
            pred, overlap = remaining_invalid.pop(0), 0.0
        else:
            pred, overlap = None, 0.0

        row_results.append(score_row(key_idx + 1, truth, pred, overlap))

    summary = build_summary(row_results, path)
    return row_results, summary


def build_summary(rows: list[dict], path: Path) -> dict:
    stem = path.stem
    group_id, _, group_name = stem.partition("_")

    interval_acc = mean(r["interval_accuracy"] for r in rows) if rows else 0.0
    category_acc = mean(r["category_accuracy"] for r in rows) if rows else 0.0
    artifact_acc = mean(r["artifact_accuracy"] for r in rows) if rows else 0.0
    general_acc  = mean([interval_acc, category_acc, artifact_acc])

    return {
        "team_file":         path.name,
        "group_id":          group_id,
        "group_name":        group_name,
        "total_points":      round(sum(r["points_earned"] for r in rows), 2),
        "interval_accuracy": round(interval_acc, 4),
        "category_accuracy": round(category_acc, 4),
        "artifact_accuracy": round(artifact_acc, 4),
        "general_accuracy":  round(general_acc,  4),
    }


# ────────────────────────────────────────────────────────────
# Output formatting
# ────────────────────────────────────────────────────────────

def format_row_report(rows: list[dict]) -> list[dict]:
    return [
        {
            "Point":             r["points_earned"],
            "Interval Accuracy": r["interval_accuracy"],
            "Category Accuracy": r["category_accuracy"],
            "Artifact Accuracy": r["artifact_accuracy"],
        }
        for r in rows
    ]


def format_summary(summary: dict) -> dict:
    return {
        "Total Points":      summary["total_points"],
        "Interval Accuracy": summary["interval_accuracy"],
        "Category Accuracy": summary["category_accuracy"],
        "Artifact Accuracy": summary["artifact_accuracy"],
        "General Accuracy":  summary["general_accuracy"],
    }


def print_summary(summary: dict) -> None:
    print("\nSubmission summary")
    print(f"  Team file          : {summary['team_file']}")
    print(f"  Group ID           : {summary['group_id']}")
    print(f"  Group name         : {summary['group_name']}")
    print(f"  Total points       : {summary['total_points']}")
    print(f"  Interval accuracy  : {summary['interval_accuracy']:.4f}")
    print(f"  Category accuracy  : {summary['category_accuracy']:.4f}")
    print(f"  Artifact accuracy  : {summary['artifact_accuracy']:.4f}")
    print(f"  General accuracy   : {summary['general_accuracy']:.4f}")


# ────────────────────────────────────────────────────────────
# CLI
# ────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Score an EEG artifact submission CSV.")
    parser.add_argument("submission", type=Path, help="Team CSV (group-id_group-name.csv)")
    parser.add_argument("--rows",    type=Path, help="Row-level output CSV")
    parser.add_argument("--summary", type=Path, help="Team summary output CSV")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base = Path("results") / args.submission.stem

    Path("results").mkdir(exist_ok=True)

    rows_path    = args.rows    or Path(f"{base}_rows.csv")
    summary_path = args.summary or Path(f"{base}_summary.csv")

    row_results, summary = score_submission(args.submission)

    write_csv(rows_path,    format_row_report(row_results))
    write_csv(summary_path, [format_summary(summary)])
    print_summary(summary)
    print(f"\nRow report   → {rows_path}") 
    print(f"Team summary → {summary_path}")


if __name__ == "__main__":
    main()  